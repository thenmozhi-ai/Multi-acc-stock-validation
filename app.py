"""
Stock Validation Dashboard
============================
Single-file Streamlit app -- everything (constants, file detection, readers,
validation logic, and Excel export) lives in this one file on purpose, so
there's no `src/` subfolder that can get lost or partially uploaded when
pushing to GitHub / Streamlit Cloud.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py
"""
from __future__ import annotations

import datetime as dt
import io
import re
import zipfile
from typing import Optional
from dataclasses import dataclass, field

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ===========================================================================
# --- from src/constants.py ---
# ===========================================================================
MARKETPLACES = ["Lazada", "Shopee", "TikTok", "Zalora"]

# File categories recognised by the detector. Each maps to a human label used in the UI.
CATEGORY_LABELS = {
    "product_master": "Product Master",
    "all_report": "ALL Report",
    "soh_report": "SOH Report",
    "mkt_lazada": "Lazada Price & Stock Report",
    "mkt_shopee": "Shopee Mass Update Report",
    "mkt_tiktok": "TikTok Batch Edit Report",
    "mkt_zalora": "Zalora Stock Report",
    "sv_lazada": "Lazada Stock Validation Report",
    "sv_shopee": "Shopee Stock Validation Report",
    "sv_tiktok": "TikTok Stock Validation Report",
    "sv_zalora": "Zalora Stock Validation Report",
}

# Remarks
REMARK_NOT_IN = "NOT IN {mkt}"
REMARK_UPDATE_0 = "UPDATE 0"
REMARK_MISMATCH = "MISMATCH STOCK"
REMARK_GOOD = "GOOD"
REMARK_REMOVE_MAX = "REMOVE MAX"

# Fill colours (hex, no leading '#') for openpyxl PatternFill, keyed by remark family
REMARK_FILL_COLORS = {
    "GOOD": "C6EFCE",
    "UPDATE 0": "FFEB9C",
    "MISMATCH STOCK": "FFC7CE",
    "NOT IN": "D9D9D9",       # prefix match for "NOT IN <MARKETPLACE>"
    "REMOVE MAX": "BDD7EE",
}

# Streamlit-side (CSS) colours, mirroring the Excel fills, for the on-screen preview
REMARK_STREAMLIT_COLORS = {
    "GOOD": "#C6EFCE",
    "UPDATE 0": "#FFEB9C",
    "MISMATCH STOCK": "#FFC7CE",
    "NOT IN": "#D9D9D9",
    "REMOVE MAX": "#BDD7EE",
}

HEADER_FILL_COLOR = "1F3864"   # navy
HEADER_FONT_COLOR = "FFFFFF"   # white
HEADER_FONT_NAME = "Arial"
HEADER_FONT_SIZE = 11
BODY_FONT_NAME = "Arial"
BODY_FONT_SIZE = 10
ALT_ROW_FILL_COLOR = "F2F2F2"

# Candidate column-name substrings (case-insensitive) used to locate columns by content
# rather than fixed position, since exports vary between runs.
COLUMN_ALIASES = {
    "sku": ["seller sku", "sellersku", "seller_sku", "sku"],
    "item_title": ["item title", "product name", "title"],
    "expected_stock": ["expected stock"],
    "master_stock": ["master stock"],
    "max_stock": ["max stock"],
    "status": ["status"],
    "quantity": ["quantity", "stock", "total"],
}

# Output workbook sheet name for the dashboard
SUMMARY_SHEET_NAME = "Summary"


# ===========================================================================
# --- from src/file_detection.py ---
# ===========================================================================
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd



@dataclass
class DetectedFiles:
    """Holds the classified uploads. Single-file categories store one UploadedFile;
    tiktok marketplace stock can legitimately be split ACTIVE/INACTIVE, so it's a list."""
    product_master: Optional[object] = None
    all_report: Optional[object] = None
    soh_report: Optional[object] = None
    mkt_lazada: Optional[object] = None
    mkt_shopee: Optional[object] = None
    mkt_tiktok: list = field(default_factory=list)
    mkt_zalora: Optional[object] = None
    mkt_zalora_status: Optional[object] = None  # optional SellerStatusTemplate
    sv_lazada: Optional[object] = None
    sv_shopee: Optional[object] = None
    sv_shopee_delist: Optional[object] = None
    sv_tiktok: Optional[object] = None
    sv_zalora: Optional[object] = None
    unrecognised: list = field(default_factory=list)

    def as_summary_rows(self) -> list[dict]:
        """Flat list of {category, label, filename} for the confirmation table."""
        rows = []

        def add(cat_key, file_obj):
            if file_obj is None:
                return
            rows.append(
                {
                    "Detected as": CATEGORY_LABELS.get(cat_key, cat_key),
                    "Filename": getattr(file_obj, "name", str(file_obj)),
                }
            )

        add("product_master", self.product_master)
        add("all_report", self.all_report)
        add("soh_report", self.soh_report)
        add("mkt_lazada", self.mkt_lazada)
        add("mkt_shopee", self.mkt_shopee)
        for f in self.mkt_tiktok:
            add("mkt_tiktok", f)
        add("mkt_zalora", self.mkt_zalora)
        if self.mkt_zalora_status is not None:
            rows.append(
                {
                    "Detected as": "Zalora Status Template (optional)",
                    "Filename": getattr(self.mkt_zalora_status, "name", "?"),
                }
            )
        add("sv_lazada", self.sv_lazada)
        add("sv_shopee", self.sv_shopee)
        if self.sv_shopee_delist is not None:
            rows.append(
                {
                    "Detected as": "Shopee Stock Validation (DELIST subset)",
                    "Filename": getattr(self.sv_shopee_delist, "name", "?"),
                }
            )
        add("sv_tiktok", self.sv_tiktok)
        add("sv_zalora", self.sv_zalora)
        return rows


def _name(f) -> str:
    return getattr(f, "name", str(f)).lower()


def detect_files(uploaded_files: list) -> DetectedFiles:
    """Classify a flat list of uploaded file objects (Streamlit UploadedFile-like:
    must expose .name) into a DetectedFiles bundle."""
    det = DetectedFiles()

    for f in uploaded_files:
        n = _name(f)

        # --- StockValidation CSVs (check these before the generic marketplace files,
        # since e.g. "stockvalidation-lazada" also contains "lazada") ---
        if "stockvalidation" in n.replace("_", "").replace(" ", ""):
            if "lazada" in n:
                det.sv_lazada = f
            elif "shopee" in n:
                if "delist" in n:
                    det.sv_shopee_delist = f
                else:
                    det.sv_shopee = f
            elif "tiktok" in n:
                det.sv_tiktok = f
            elif "zalora" in n:
                det.sv_zalora = f
            else:
                det.unrecognised.append(f)
            continue

        # --- Marketplace stock/status files ---
        if "pricestock" in n:
            det.mkt_lazada = f
            continue
        if "mass_update_sales_info" in n or "massupdatesalesinfo" in n:
            det.mkt_shopee = f
            continue
        if "batchedit" in n:
            det.mkt_tiktok.append(f)
            continue
        if "sellerstocktemplate" in n:
            det.mkt_zalora = f
            continue
        if "sellerstatustemplate" in n:
            det.mkt_zalora_status = f
            continue

        # --- Reference / master files ---
        if "productmaster" in n.replace(" ", "").replace("_", ""):
            det.product_master = f
            continue
        if "soh" in n:
            det.soh_report = f
            continue
        if n.startswith("all") or n.startswith("all_") or n.startswith("all-") or n.startswith("all "):
            det.all_report = f
            continue

        det.unrecognised.append(f)

    return det


def sniff_columns(f) -> list[str]:
    """Best-effort peek at a file's header row, used only for diagnostics /
    disambiguation messages shown to the user -- never required for detection to work."""
    try:
        name = _name(f)
        f.seek(0)
        if name.endswith(".csv"):
            df = pd.read_csv(f, nrows=0)
        else:
            df = pd.read_excel(f, nrows=0)
        f.seek(0)
        return list(df.columns)
    except Exception:
        try:
            f.seek(0)
        except Exception:
            pass
        return []


# ===========================================================================
# --- from src/readers.py ---
# ===========================================================================
import io
import re
import zipfile
from typing import Optional

import pandas as pd



# --------------------------------------------------------------------------- #
# Column-matching helpers
# --------------------------------------------------------------------------- #

def _find_column(df: pd.DataFrame, aliases: list[str]) -> Optional[str]:
    """Return the first column in df whose name contains any of the given
    case-insensitive substrings, longest alias first (so 'seller sku' wins over 'sku')."""
    cols = list(df.columns)
    for alias in sorted(aliases, key=len, reverse=True):
        for c in cols:
            if alias in str(c).strip().lower():
                return c
    return None


def find_col(df: pd.DataFrame, key: str) -> Optional[str]:
    """Look up a column by semantic key (see COLUMN_ALIASES)."""
    return _find_column(df, COLUMN_ALIASES.get(key, [key]))


def normalize_sku_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip()


def clean_skus(df: pd.DataFrame, sku_col: str) -> pd.DataFrame:
    df = df.copy()
    df[sku_col] = normalize_sku_series(df[sku_col])
    df = df[df[sku_col].notna() & (df[sku_col] != "") & (df[sku_col].str.lower() != "nan")]
    return df


def is_bundle_sku(sku: str) -> bool:
    """Shopee combo/bundle SKUs contain a '+' and don't map 1:1 to a single SKU."""
    return "+" in str(sku)


# --------------------------------------------------------------------------- #
# Shopee activePane bug patch
# --------------------------------------------------------------------------- #

def _patch_shopee_workbook(file_bytes: bytes) -> io.BytesIO:
    """
    Some Shopee mass_update_sales_info exports ship with an invalid
    `activePane` attribute inside xl/worksheets/sheetN.xml's <pane> element,
    which makes openpyxl/pandas choke on load. Patch it out in-memory by
    rewriting the offending attribute before parsing.
    """
    try:
        src = zipfile.ZipFile(io.BytesIO(file_bytes), "r")
    except zipfile.BadZipFile:
        # Not actually a zip-based xlsx (e.g. already CSV-like) -- return as-is.
        return io.BytesIO(file_bytes)

    out_buffer = io.BytesIO()
    with zipfile.ZipFile(out_buffer, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8", errors="ignore")
                # Strip an invalid activePane reference (e.g. activePane="topLeft" combined
                # with a malformed pane split) that some Shopee exports emit.
                text = re.sub(r'\s*activePane="[^"]*"', "", text)
                data = text.encode("utf-8")
            dst.writestr(item, data)
    src.close()
    out_buffer.seek(0)
    return out_buffer


def _read_excel_any(f, **kwargs) -> pd.DataFrame:
    """Read an uploaded file as Excel regardless of whether it arrives as a
    Streamlit UploadedFile, file path, or raw bytes buffer."""
    if hasattr(f, "read"):
        f.seek(0)
        data = f.read()
        f.seek(0)
    else:
        data = f
    return pd.read_excel(io.BytesIO(data), **kwargs)


# --------------------------------------------------------------------------- #
# StockValidation CSV reader (shared shape across all 4 marketplaces)
# --------------------------------------------------------------------------- #

def read_stock_validation_csv(f) -> pd.DataFrame:
    """
    Reads a stockValidation-<marketplace>.csv file. Expected (flexibly-matched)
    columns: SKU / Seller SKU, Item Title, Expected Stock, Max Stock (optional),
    Status (optional), Master Stock (optional).
    """
    if hasattr(f, "seek"):
        f.seek(0)
    df = pd.read_csv(f)

    sku_col = find_col(df, "sku")
    if sku_col is None:
        raise ValueError("Could not find a SKU / Seller SKU column in this StockValidation file.")
    exp_col = find_col(df, "expected_stock")
    if exp_col is None:
        raise ValueError("Could not find an 'Expected Stock' column in this StockValidation file.")

    out = pd.DataFrame()
    out["SKU"] = df[sku_col]
    out["Expected Stock"] = pd.to_numeric(df[exp_col], errors="coerce").fillna(0)

    title_col = find_col(df, "item_title")
    out["Item Title"] = df[title_col] if title_col else ""

    max_col = find_col(df, "max_stock")
    out["Max Stock"] = pd.to_numeric(df[max_col], errors="coerce") if max_col else pd.NA

    status_col = find_col(df, "status")
    out["Status"] = df[status_col] if status_col else ""

    master_col = find_col(df, "master_stock")
    if master_col:
        out["Master Stock"] = pd.to_numeric(df[master_col], errors="coerce")

    out = clean_skus(out, "SKU")
    return out


# --------------------------------------------------------------------------- #
# Marketplace stock readers
# --------------------------------------------------------------------------- #

def read_lazada_pricestock(f) -> pd.DataFrame:
    """Header row 0, skip rows 1-3. Column count can vary (15 vs 16 -- some exports
    add a Barcode column) so columns are matched by name, not position."""
    df = _read_excel_any(f, header=0, skiprows=[1, 2, 3])
    sku_col = find_col(df, "sku")
    qty_col = find_col(df, "quantity")
    status_col = find_col(df, "status")
    if sku_col is None or qty_col is None:
        raise ValueError("Could not locate SellerSKU / Quantity columns in the Lazada pricestock file.")
    out = pd.DataFrame()
    out["SKU"] = df[sku_col]
    out["Marketplace Stock"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    out["Marketplace Status"] = df[status_col] if status_col else ""
    out = clean_skus(out, "SKU")
    out = out.groupby("SKU", as_index=False).agg(
        {"Marketplace Stock": "sum", "Marketplace Status": "first"}
    )
    return out


def read_shopee_mass_update(f) -> pd.DataFrame:
    """Patches the activePane bug, then header row 2, skip rows 3-5.
    Matches SKU, falling back to Parent SKU when SKU is blank."""
    if hasattr(f, "seek"):
        f.seek(0)
    data = f.read() if hasattr(f, "read") else f
    patched = _patch_shopee_workbook(data)
    df = pd.read_excel(patched, header=2, skiprows=[3, 4, 5])

    sku_col = _find_column(df, ["sku"])
    parent_col = _find_column(df, ["parent sku"])
    qty_col = find_col(df, "quantity")
    if sku_col is None or qty_col is None:
        raise ValueError("Could not locate SKU / Stock columns in the Shopee mass_update file.")

    out = pd.DataFrame()
    sku_series = df[sku_col]
    if parent_col is not None:
        sku_series = sku_series.where(sku_series.notna() & (sku_series.astype(str).str.strip() != ""), df[parent_col])
    out["SKU"] = sku_series
    out["Marketplace Stock"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    out = clean_skus(out, "SKU")

    # Exclude bundle SKUs (contain '+') -- combo listings, not 1:1 with a single SKU.
    out = out[~out["SKU"].apply(is_bundle_sku)]
    out = out.groupby("SKU", as_index=False).agg({"Marketplace Stock": "sum"})
    return out


def read_shopee_delist(f) -> set:
    """Reads a Shopee _DELIST export and returns the set of SKUs it contains
    (used to mark those SKUs INACTIVE relative to the full export)."""
    if hasattr(f, "seek"):
        f.seek(0)
    data = f.read() if hasattr(f, "read") else f
    patched = _patch_shopee_workbook(data)
    df = pd.read_excel(patched, header=2, skiprows=[3, 4, 5])
    sku_col = _find_column(df, ["sku"])
    if sku_col is None:
        return set()
    skus = normalize_sku_series(df[sku_col])
    return set(skus[skus.notna() & (skus != "")])


def read_tiktok_batchedit_single(f) -> pd.DataFrame:
    """Header row 2, skip rows 3-4."""
    df = _read_excel_any(f, header=2, skiprows=[3, 4])
    sku_col = find_col(df, "sku")
    qty_col = find_col(df, "quantity")
    if sku_col is None or qty_col is None:
        raise ValueError("Could not locate Seller SKU / Quantity columns in the TikTok batchedit file.")
    out = pd.DataFrame()
    out["SKU"] = df[sku_col]
    out["Marketplace Stock"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    out = clean_skus(out, "SKU")
    return out


def read_tiktok_batchedit(files: list) -> pd.DataFrame:
    """Accepts one or two TikTok batchedit files (a single combined export, or a
    split ACTIVE + INACTIVE pair). If two, concatenates and sums duplicate SKUs."""
    frames = [read_tiktok_batchedit_single(f) for f in files]
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=["SKU", "Marketplace Stock"])
    if combined.empty:
        return combined
    combined = combined.groupby("SKU", as_index=False).agg({"Marketplace Stock": "sum"})
    return combined


def read_zalora_stock(f, status_f=None) -> pd.DataFrame:
    """Header row 0. Optional SellerStatusTemplate file appends Zalora_Status."""
    df = _read_excel_any(f, header=0)
    sku_col = find_col(df, "sku")
    qty_col = find_col(df, "quantity")
    if sku_col is None or qty_col is None:
        raise ValueError("Could not locate SellerSku / Quantity columns in the Zalora stock file.")
    out = pd.DataFrame()
    out["SKU"] = df[sku_col]
    out["Marketplace Stock"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    out = clean_skus(out, "SKU")

    if status_f is not None:
        try:
            sdf = _read_excel_any(status_f, header=0)
            s_sku_col = find_col(sdf, "sku")
            s_status_col = find_col(sdf, "status")
            if s_sku_col and s_status_col:
                sdf2 = pd.DataFrame(
                    {"SKU": normalize_sku_series(sdf[s_sku_col]), "Marketplace Status": sdf[s_status_col]}
                )
                out = out.merge(sdf2, on="SKU", how="left")
        except Exception:
            pass

    if "Marketplace Status" not in out.columns:
        out["Marketplace Status"] = ""

    out = out.groupby("SKU", as_index=False).agg(
        {"Marketplace Stock": "sum", "Marketplace Status": "first"}
    )
    return out


# --------------------------------------------------------------------------- #
# Reference file readers (Product Master / ALL report / SOH) -- optional,
# used only as a presence-check cross-reference against Expected Stock.
# --------------------------------------------------------------------------- #

def read_reference_file(f) -> pd.DataFrame:
    """Reads a Product Master / ALL report / SOH file generically: whatever SKU +
    quantity-like columns it can find. Used only for an optional presence-check,
    so failures here are non-fatal."""
    try:
        name = getattr(f, "name", "").lower()
        if hasattr(f, "seek"):
            f.seek(0)
        if name.endswith(".csv"):
            df = pd.read_csv(f)
        else:
            df = _read_excel_any(f)
        sku_col = find_col(df, "sku")
        if sku_col is None:
            return pd.DataFrame(columns=["SKU"])
        out = pd.DataFrame()
        out["SKU"] = normalize_sku_series(df[sku_col])
        out = out[out["SKU"].notna() & (out["SKU"] != "")]
        return out.drop_duplicates(subset="SKU")
    except Exception:
        return pd.DataFrame(columns=["SKU"])


# ===========================================================================
# --- from src/validation.py ---
# ===========================================================================
import pandas as pd



def cross_check_reference(sv_df: pd.DataFrame, reference_df: pd.DataFrame) -> dict:
    """Presence-check only: what fraction of StockValidation SKUs are also found in
    the optional Product Master / SOH reference file. Does NOT change Expected Stock --
    the StockValidation CSV's own 'Expected Stock' column is already correctly derived
    downstream, per the underlying skill."""
    if reference_df is None or reference_df.empty or sv_df.empty:
        return {"checked": False}
    ref_skus = set(reference_df["SKU"])
    sv_skus = set(sv_df["SKU"])
    found = sv_skus & ref_skus
    return {
        "checked": True,
        "sv_sku_count": len(sv_skus),
        "found_in_reference": len(found),
        "match_rate": (len(found) / len(sv_skus)) if sv_skus else 0.0,
    }


def validate_marketplace(
    marketplace: str,
    sv_df: pd.DataFrame,
    mkt_df: pd.DataFrame,
    flag_max_zero: bool = False,
) -> pd.DataFrame:
    """
    Joins a marketplace's StockValidation export against its own live-stock export
    and assigns a remark per SKU, in priority order:

      1. SKU not found in the marketplace's own stock export -> NOT IN <MARKETPLACE>
      2. (optional) Max Stock == 0 and (Expected or Marketplace stock) > 0 -> REMOVE MAX
      3. Expected Stock == 0 and Marketplace Stock > 0 -> UPDATE 0
      4. Expected Stock != Marketplace Stock -> MISMATCH STOCK
      5. otherwise -> GOOD
    """
    if sv_df is None or sv_df.empty:
        return pd.DataFrame()

    merged = sv_df.merge(mkt_df, on="SKU", how="left", suffixes=("", "_mkt"))
    if "Marketplace Stock" not in merged.columns:
        merged["Marketplace Stock"] = pd.NA
    if "Marketplace Status" not in merged.columns:
        merged["Marketplace Status"] = ""

    def remark_row(row):
        if pd.isna(row["Marketplace Stock"]):
            return REMARK_NOT_IN.format(mkt=marketplace.upper())

        expected = row.get("Expected Stock", 0) or 0
        mkt_stock = row["Marketplace Stock"] or 0

        if flag_max_zero:
            max_stock = row.get("Max Stock")
            if pd.notna(max_stock) and max_stock == 0 and (expected > 0 or mkt_stock > 0):
                return REMARK_REMOVE_MAX

        if expected == 0 and mkt_stock > 0:
            return REMARK_UPDATE_0
        if expected != mkt_stock:
            return REMARK_MISMATCH
        return REMARK_GOOD

    merged["Difference"] = merged["Marketplace Stock"].fillna(0) - merged.get("Expected Stock", 0)
    merged["Remark"] = merged.apply(remark_row, axis=1)
    merged["Marketplace"] = marketplace

    cols = [
        "Marketplace",
        "SKU",
        "Item Title",
        "Expected Stock",
        "Marketplace Stock",
        "Difference",
    ]
    if "Max Stock" in merged.columns:
        cols.append("Max Stock")
    if "Marketplace Status" in merged.columns:
        cols.append("Marketplace Status")
    cols.append("Remark")
    cols = [c for c in cols if c in merged.columns]
    return merged[cols]


def apply_shopee_delist(mkt_df: pd.DataFrame, delist_skus: set) -> pd.DataFrame:
    """Marks SKUs found in the Shopee _DELIST export as INACTIVE, everything else ACTIVE."""
    if not delist_skus or mkt_df.empty:
        return mkt_df
    mkt_df = mkt_df.copy()
    mkt_df["Marketplace Status"] = mkt_df["SKU"].apply(
        lambda s: "INACTIVE" if s in delist_skus else "ACTIVE"
    )
    return mkt_df


def summarize(results: dict) -> pd.DataFrame:
    """Builds the KPI summary table (one row per marketplace) from the
    {marketplace: result_df} dict produced by validate_marketplace calls."""
    rows = []
    for mkt, df in results.items():
        if df is None or df.empty:
            continue
        total = len(df)
        good = (df["Remark"] == REMARK_GOOD).sum()
        mismatch = (df["Remark"] == REMARK_MISMATCH).sum()
        update0 = (df["Remark"] == REMARK_UPDATE_0).sum()
        remove_max = (df["Remark"] == REMARK_REMOVE_MAX).sum()
        not_in = df["Remark"].str.startswith("NOT IN").sum()
        issues = total - good
        accuracy = (good / total * 100) if total else 0.0
        rows.append(
            {
                "Marketplace": mkt,
                "Total SKUs": total,
                "GOOD": int(good),
                "MISMATCH STOCK": int(mismatch),
                "UPDATE 0": int(update0),
                "REMOVE MAX": int(remove_max),
                "NOT IN MARKETPLACE": int(not_in),
                "Total Issues": int(issues),
                "Accuracy %": round(accuracy, 1),
            }
        )
    return pd.DataFrame(rows)


# ===========================================================================
# --- from src/excel_export.py ---
# ===========================================================================
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


THIN = Side(style="thin", color="B7B7B7")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws: Worksheet, row: int, ncols: int):
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(name=HEADER_FONT_NAME, size=HEADER_FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _remark_fill_for(remark: str) -> str | None:
    if remark is None:
        return None
    remark = str(remark)
    if remark in REMARK_FILL_COLORS:
        return REMARK_FILL_COLORS[remark]
    if remark.startswith("NOT IN"):
        return REMARK_FILL_COLORS["NOT IN"]
    return None


def _write_dataframe(ws: Worksheet, df: pd.DataFrame, start_row: int = 1, remark_col_name: str | None = "Remark"):
    body_font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_FILL_COLOR)

    # Header
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col))
    _style_header(ws, start_row, len(df.columns))

    remark_idx = list(df.columns).index(remark_col_name) + 1 if remark_col_name in df.columns else None

    # Body
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        r = start_row + i
        remark_val = row.get(remark_col_name) if remark_col_name else None
        remark_fill_color = _remark_fill_for(remark_val) if remark_val is not None else None
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = body_font
            cell.border = THIN_BORDER
            if remark_fill_color:
                cell.fill = PatternFill("solid", fgColor=remark_fill_color)
            elif i % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    for j, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 10), 40)

    # Freeze header + autofilter
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    last_col_letter = get_column_letter(len(df.columns))
    ws.auto_filter.ref = f"A{start_row}:{last_col_letter}{start_row + len(df)}"


def build_workbook(summary_df: pd.DataFrame, marketplace_results: dict, run_meta: dict | None = None) -> bytes:
    """
    summary_df: output of validation.summarize()
    marketplace_results: {marketplace_name: result_df} from validation.validate_marketplace()
    run_meta: optional dict of extra key/value lines to print above the summary table
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = SUMMARY_SHEET_NAME

    row_cursor = 1
    title_font = Font(name=HEADER_FONT_NAME, size=14, bold=True, color=HEADER_FILL_COLOR)
    ws_summary.cell(row=row_cursor, column=1, value="Stock Validation Summary").font = title_font
    row_cursor += 1

    if run_meta:
        meta_font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE, italic=True)
        for k, v in run_meta.items():
            ws_summary.cell(row=row_cursor, column=1, value=f"{k}: {v}").font = meta_font
            row_cursor += 1

    row_cursor += 1  # blank row
    if not summary_df.empty:
        _write_dataframe(ws_summary, summary_df, start_row=row_cursor, remark_col_name=None)
    else:
        ws_summary.cell(row=row_cursor, column=1, value="No marketplaces had both a StockValidation file and a stock file uploaded.")

    # One sheet per marketplace
    for mkt, df in marketplace_results.items():
        if df is None or df.empty:
            continue
        ws = wb.create_sheet(title=mkt[:31])
        export_df = df.drop(columns=["Marketplace"]) if "Marketplace" in df.columns else df
        _write_dataframe(ws, export_df, start_row=1, remark_col_name="Remark")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================================
# --- Streamlit UI ---
# ===========================================================================

st.set_page_config(page_title="Stock Validation Dashboard", page_icon="📦", layout="wide")

ACCOUNTS = ["DBC", "EWG", "IEI"]
MARKETPLACES_UI = ["Lazada", "Shopee", "TikTok", "Zalora"]

# Which reference files each account uses. EWG has no Warehouse SOH.
ACCOUNT_REFERENCE_FILES = {
    "IEI": ["Product Master", "SOH Report", "ALL Report"],
    "DBC": ["Product Master", "SOH Report", "ALL Report"],
    "EWG": ["Product Master", "ALL Report"],
}

MKT_STOCK_LABEL = {
    "Lazada": "Lazada Price & Stock Report",
    "Shopee": "Shopee Mass Update Report",
    "TikTok": "TikTok Batch Edit Report",
    "Zalora": "Zalora Stock Report",
}
MKT_SV_LABEL = {
    "Lazada": "Lazada Stock Validation Report",
    "Shopee": "Shopee Stock Validation Report",
    "TikTok": "TikTok Stock Validation Report",
    "Zalora": "Zalora Stock Validation Report",
}
ACCOUNT_NOTES = {
    "IEI": "Uses SOH + Product Master/ALL Report as an optional cross-check. "
    "Expected Stock always comes from each marketplace's own Stock Validation Report.",
    "DBC": "Same file logic as IEI. (Listing-status recommendations, e.g. "
    "TikTok-driven Active/Inactive, aren't produced by this dashboard -- it's stock-only.)",
    "EWG": "No Warehouse SOH for this account -- Expected Stock is cross-checked "
    "against the Product Master / ALL Report instead.",
}


# --------------------------------------------------------------------------- #
# Sidebar: account selector + run controls
# --------------------------------------------------------------------------- #

st.sidebar.title("📦 Stock Validation")

account = st.sidebar.selectbox("Account", ACCOUNTS, index=0, key="account_select")
st.sidebar.caption(ACCOUNT_NOTES[account])

st.sidebar.divider()

flag_max_zero = st.sidebar.checkbox(
    "Flag REMOVE MAX (Max Stock = 0 while real stock exists)",
    value=(account == "EWG"),
    help="Rows where the Stock Validation Report's 'Max Stock' column is 0 but "
    "Expected or Marketplace stock is > 0 get a REMOVE MAX remark instead of the "
    "usual mismatch remark -- this catches listing caps silently blocking sales.",
)

run_clicked = st.sidebar.button("▶ Run Validation", type="primary", width='stretch')


# --------------------------------------------------------------------------- #
# Header
# --------------------------------------------------------------------------- #

st.title("📦 Stock Validation Dashboard")
st.caption(
    f"Account: **{account}** -- reconciles Expected Stock against live marketplace "
    "stock for Lazada, Shopee, TikTok, and Zalora, flags mismatches, and exports a "
    "formatted Excel workbook."
)

st.subheader(f"1. Upload files for {account}")

# --- Reference files ---
ref_categories = ACCOUNT_REFERENCE_FILES[account]
with st.expander("Reference Files (optional)", expanded=True):
    ref_cols = st.columns(len(ref_categories))
    ref_uploads = {}
    for col, label in zip(ref_cols, ref_categories):
        with col:
            ref_uploads[label] = st.file_uploader(
                label, type=["csv", "xlsx", "xls"], key=f"ref_{label}"
            )

# --- Per-marketplace files ---
st.markdown("**Marketplace Files**")
mkt_stock_uploads = {}
mkt_sv_uploads = {}
mkt_extra_uploads = {}  # Zalora status template, kept optional

for mkt in MARKETPLACES_UI:
    with st.expander(mkt, expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            multi = mkt in ("Shopee", "TikTok")
            help_txt = None
            if mkt == "Shopee":
                help_txt = "Upload the main export. If you also have a separate _DELIST export, add both here."
            elif mkt == "TikTok":
                help_txt = "Upload one combined export, or both ACTIVE + INACTIVE files if yours are split."
            mkt_stock_uploads[mkt] = st.file_uploader(
                MKT_STOCK_LABEL[mkt],
                type=["csv", "xlsx", "xls"],
                accept_multiple_files=multi,
                help=help_txt,
                key=f"mktstock_{mkt}",
            )
        with c2:
            mkt_sv_uploads[mkt] = st.file_uploader(
                MKT_SV_LABEL[mkt], type=["csv"], key=f"sv_{mkt}"
            )
        if mkt == "Zalora":
            mkt_extra_uploads["Zalora_status"] = st.file_uploader(
                "Zalora Status Template (optional)",
                type=["csv", "xlsx", "xls"],
                key="zalora_status",
            )


def _has_file(x) -> bool:
    if x is None:
        return False
    if isinstance(x, list):
        return len(x) > 0
    return True


marketplaces_ready = [
    mkt for mkt in MARKETPLACES_UI
    if _has_file(mkt_stock_uploads.get(mkt)) and _has_file(mkt_sv_uploads.get(mkt))
]

if marketplaces_ready:
    st.success(f"Ready to validate: **{', '.join(marketplaces_ready)}**")
else:
    st.info(
        "Upload both the stock file and the Stock Validation Report for at least "
        "one marketplace above, then click **▶ Run Validation** in the sidebar."
    )

if not run_clicked:
    st.stop()

if not marketplaces_ready:
    st.error("Can't run validation yet -- see the notice above.")
    st.stop()


# --------------------------------------------------------------------------- #
# Run validation
# --------------------------------------------------------------------------- #

with st.spinner("Reading files and validating stock..."):
    errors = []
    reference_frames = []

    for label, f in ref_uploads.items():
        if f is not None:
            try:
                reference_frames.append(read_reference_file(f))
            except Exception as e:
                errors.append(f"{label}: {e}")

    reference_df = (
        pd.concat(reference_frames, ignore_index=True).drop_duplicates(subset="SKU")
        if reference_frames
        else pd.DataFrame(columns=["SKU"])
    )

    results = {}
    cross_checks = {}

    # Lazada
    if "Lazada" in marketplaces_ready:
        try:
            sv = read_stock_validation_csv(mkt_sv_uploads["Lazada"])
            mkt_df = read_lazada_pricestock(mkt_stock_uploads["Lazada"])
            cross_checks["Lazada"] = cross_check_reference(sv, reference_df)
            results["Lazada"] = validate_marketplace("Lazada", sv, mkt_df, flag_max_zero)
        except Exception as e:
            errors.append(f"Lazada: {e}")

    # Shopee (may have 1 or 2 files: full export + optional _DELIST export)
    if "Shopee" in marketplaces_ready:
        try:
            shopee_files = mkt_stock_uploads["Shopee"]
            shopee_files = shopee_files if isinstance(shopee_files, list) else [shopee_files]
            delist_file = next((f for f in shopee_files if "delist" in f.name.lower()), None)
            full_file = next((f for f in shopee_files if f is not delist_file), shopee_files[0])

            sv = read_stock_validation_csv(mkt_sv_uploads["Shopee"])
            mkt_df = read_shopee_mass_update(full_file)
            if delist_file is not None:
                delist_skus = read_shopee_delist(delist_file)
                mkt_df = apply_shopee_delist(mkt_df, delist_skus)
            cross_checks["Shopee"] = cross_check_reference(sv, reference_df)
            results["Shopee"] = validate_marketplace("Shopee", sv, mkt_df, flag_max_zero)
        except Exception as e:
            errors.append(f"Shopee: {e}")

    # TikTok (may have 1 combined file, or 2: ACTIVE + INACTIVE)
    if "TikTok" in marketplaces_ready:
        try:
            tiktok_files = mkt_stock_uploads["TikTok"]
            tiktok_files = tiktok_files if isinstance(tiktok_files, list) else [tiktok_files]
            sv = read_stock_validation_csv(mkt_sv_uploads["TikTok"])
            mkt_df = read_tiktok_batchedit(tiktok_files)
            cross_checks["TikTok"] = cross_check_reference(sv, reference_df)
            results["TikTok"] = validate_marketplace("TikTok", sv, mkt_df, flag_max_zero)
        except Exception as e:
            errors.append(f"TikTok: {e}")

    # Zalora
    if "Zalora" in marketplaces_ready:
        try:
            sv = read_stock_validation_csv(mkt_sv_uploads["Zalora"])
            mkt_df = read_zalora_stock(mkt_stock_uploads["Zalora"], mkt_extra_uploads.get("Zalora_status"))
            cross_checks["Zalora"] = cross_check_reference(sv, reference_df)
            results["Zalora"] = validate_marketplace("Zalora", sv, mkt_df, flag_max_zero)
        except Exception as e:
            errors.append(f"Zalora: {e}")

if errors:
    st.error("Some marketplaces couldn't be processed:")
    for e in errors:
        st.write(f"- {e}")

if not results:
    st.stop()

summary_df = summarize(results)

# --------------------------------------------------------------------------- #
# Dashboard
# --------------------------------------------------------------------------- #

st.subheader("2. Summary dashboard")

total_skus = int(summary_df["Total SKUs"].sum()) if not summary_df.empty else 0
total_issues = int(summary_df["Total Issues"].sum()) if not summary_df.empty else 0
overall_accuracy = (
    round((summary_df["GOOD"].sum() / total_skus) * 100, 1) if total_skus else 0.0
)

kpi_cols = st.columns(4)
kpi_cols[0].metric("Total SKUs validated", f"{total_skus:,}")
kpi_cols[1].metric("Marketplaces covered", len(results))
kpi_cols[2].metric("Total issues flagged", f"{total_issues:,}")
kpi_cols[3].metric("Overall accuracy", f"{overall_accuracy}%")

if reference_frames:
    with st.expander("Reference file cross-check (Product Master / ALL / SOH presence-check)"):
        for mkt, cc in cross_checks.items():
            if cc.get("checked"):
                st.write(
                    f"**{mkt}**: {cc['found_in_reference']}/{cc['sv_sku_count']} SKUs found "
                    f"in the uploaded reference file(s) ({cc['match_rate']*100:.1f}%)."
                )

st.dataframe(summary_df, width='stretch', hide_index=True)

if not summary_df.empty:
    chart_df = summary_df.set_index("Marketplace")[
        ["GOOD", "MISMATCH STOCK", "UPDATE 0", "REMOVE MAX", "NOT IN MARKETPLACE"]
    ]
    st.bar_chart(chart_df)


# --------------------------------------------------------------------------- #
# Per-marketplace detail tabs (doubles as the Mismatch Report -- filter to any
# non-GOOD remark to see just the mismatches)
# --------------------------------------------------------------------------- #

st.subheader("3. Marketplace detail & mismatch report")


def _style_remark(val: str) -> str:
    color = REMARK_STREAMLIT_COLORS.get(val)
    if not color and isinstance(val, str) and val.startswith("NOT IN"):
        color = REMARK_STREAMLIT_COLORS["NOT IN"]
    return f"background-color: {color}" if color else ""


tabs = st.tabs(list(results.keys()))
for tab, mkt in zip(tabs, results.keys()):
    with tab:
        df = results[mkt]
        remark_filter = st.multiselect(
            f"Filter {mkt} remarks",
            options=sorted(df["Remark"].unique().tolist()),
            default=[],
            key=f"filter_{mkt}",
        )
        view_df = df[df["Remark"].isin(remark_filter)] if remark_filter else df
        display_cols = [c for c in view_df.columns if c != "Marketplace"]
        styled = view_df[display_cols].style.map(_style_remark, subset=["Remark"])
        st.dataframe(styled, width='stretch', hide_index=True)


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #

st.subheader("4. Export")

run_meta = {
    "Account": account,
    "Generated": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
    "Marketplaces": ", ".join(results.keys()),
    "REMOVE MAX flagging": "On" if flag_max_zero else "Off",
}
workbook_bytes = build_workbook(summary_df, results, run_meta)

st.download_button(
    label="⬇ Download Excel workbook",
    data=workbook_bytes,
    file_name=f"{account}_Stock_Validation_{dt.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    width='stretch',
)
