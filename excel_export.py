"""Builds the final colour-coded Excel workbook: Summary sheet + one sheet per
marketplace actually validated, styled to match the underlying stock-validation skill."""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.constants import (
    ALT_ROW_FILL_COLOR,
    BODY_FONT_NAME,
    BODY_FONT_SIZE,
    HEADER_FILL_COLOR,
    HEADER_FONT_COLOR,
    HEADER_FONT_NAME,
    HEADER_FONT_SIZE,
    REMARK_FILL_COLORS,
    SUMMARY_SHEET_NAME,
)

THIN = Side(style="thin", color="B7B7B7")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws: Worksheet, row: int, ncols: int):
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(name=HEADER_FONT_NAME, size=HEADER_FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _remark_fill_for(remark: str) -> str | None:
    if remark is None:
        return None
    remark = str(remark)
    if remark in REMARK_FILL_COLORS:
        return REMARK_FILL_COLORS[remark]
    if remark.startswith("NOT IN"):
        return REMARK_FILL_COLORS["NOT IN"]
    return None


def _write_dataframe(ws: Worksheet, df: pd.DataFrame, start_row: int = 1, remark_col_name: str | None = "Remark"):
    body_font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_FILL_COLOR)

    # Header
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col))
    _style_header(ws, start_row, len(df.columns))

    remark_idx = list(df.columns).index(remark_col_name) + 1 if remark_col_name in df.columns else None

    # Body
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        r = start_row + i
        remark_val = row.get(remark_col_name) if remark_col_name else None
        remark_fill_color = _remark_fill_for(remark_val) if remark_val is not None else None
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = body_font
            cell.border = THIN_BORDER
            if remark_fill_color:
                cell.fill = PatternFill("solid", fgColor=remark_fill_color)
            elif i % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    for j, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 10), 40)

    # Freeze header + autofilter
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    last_col_letter = get_column_letter(len(df.columns))
    ws.auto_filter.ref = f"A{start_row}:{last_col_letter}{start_row + len(df)}"


def build_workbook(summary_df: pd.DataFrame, marketplace_results: dict, run_meta: dict | None = None) -> bytes:
    """
    summary_df: output of validation.summarize()
    marketplace_results: {marketplace_name: result_df} from validation.validate_marketplace()
    run_meta: optional dict of extra key/value lines to print above the summary table
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = SUMMARY_SHEET_NAME

    row_cursor = 1
    title_font = Font(name=HEADER_FONT_NAME, size=14, bold=True, color=HEADER_FILL_COLOR)
    ws_summary.cell(row=row_cursor, column=1, value="Stock Validation Summary").font = title_font
    row_cursor += 1

    if run_meta:
        meta_font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE, italic=True)
        for k, v in run_meta.items():
            ws_summary.cell(row=row_cursor, column=1, value=f"{k}: {v}").font = meta_font
            row_cursor += 1

    row_cursor += 1  # blank row
    if not summary_df.empty:
        _write_dataframe(ws_summary, summary_df, start_row=row_cursor, remark_col_name=None)
    else:
        ws_summary.cell(row=row_cursor, column=1, value="No marketplaces had both a StockValidation file and a stock file uploaded.")

    # One sheet per marketplace
    for mkt, df in marketplace_results.items():
        if df is None or df.empty:
            continue
        ws = wb.create_sheet(title=mkt[:31])
        export_df = df.drop(columns=["Marketplace"]) if "Marketplace" in df.columns else df
        _write_dataframe(ws, export_df, start_row=1, remark_col_name="Remark")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
